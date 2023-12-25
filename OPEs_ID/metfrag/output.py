import dataclasses
from io import BytesIO
from itertools import chain
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.units import pixels_to_points
from rdkit import Chem
from rdkit.Chem import Draw

from .param import MetFragParameter


def mol_to_excel_image(mol, size=(150, 150)):
    pic = Draw.MolToImage(mol, size=size)
    fp = BytesIO()
    pic.save(fp, format="png")
    return ExcelImage(fp)


def to_excel_with_mol_img(df: pd.DataFrame, fname: str | Path, img_size=(150, 150), metfrag_param: MetFragParameter = None):
    wb = Workbook()
    ws = wb.create_sheet('RESULTS', 0)

    ws.append(["Index", "Mol", *df.columns.to_list()])

    for row, s in enumerate(df.itertuples(), start=2):
        ws.append((s.Index, None, *s[1:]))
        try:
            img = mol_to_excel_image(Chem.MolFromInchi(s.InChI), size=img_size)
            ws.add_image(img, f"B{row}")
        except:
            pass

    ws.column_dimensions['B'].width = img_size[1] / 7
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = pixels_to_points(img_size[0])
    for c in chain(ws[1], ws['A']):
        c.style = 'Pandas'

    if metfrag_param is not None:
        ws = wb.create_sheet('PARAMS', 1)
        ws.append(['Name', 'Value'])
        for field in dataclasses.fields(metfrag_param):
            ws.append([field.name, str(getattr(metfrag_param, field.name))])
        for c in ws[1]:
            c.style = 'Pandas'

    wb.save(fname)
